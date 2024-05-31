import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def adjust_column_width_from_df(worksheet):
    '''
    Adjusts the column width of the specified Excel worksheet based on the maximum length of the data contained in the cells of each column.

    Args:
        worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object whose column width needs to be adjusted.

    Returns:
        None

    This function loops through all columns of the worksheet, scans the cells of each column to find the maximum character length, and sets the column width based on that length.
    The column width is set to the maximum character length plus 10.
    '''
    for col_index, col in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in col:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
        adjusted_width = max_length + 10
        worksheet.column_dimensions[get_column_letter(col_index)].width = adjusted_width

def export_data(df, table_name, formatted_date):
    '''
    Exports the given DataFrame to an Excel file and a CSV file. If the Excel file already exists, a new sheet is added to the existing file.
    If the file does not exist, a new file is created. At the same time, a CSV file is also exported.

    Args:
        df (pandas.DataFrame): DataFrame containing the data to be exported.
        table_name (str): The name of the table for the output data. Used as the sheet name in Excel and the filename for the CSV file.
        formatted_date (str): Date string used in the output filenames. It is incorporated as part of the filename.

    Returns:
        None

    Process flow:
        1. Check for the existence of the Excel file, load it if it exists, or create a new one if it does not.
        2. Write the DataFrame to a new sheet in the Excel file.
        3. After writing, call the column width adjustment function.
        4. Write the DataFrame to a CSV file and save it with the specified encoding.

    Exceptions:
        FileNotFoundError: Raised if the specified Excel file is not found.
    '''
    output_excel_filename = f'output/output_data_{formatted_date}.xlsx'
    output_csv_filename = f'output/{table_name}_output_data_{formatted_date}.csv'
    
    try:
        # Check if the Excel file already exists and load it if it does
        book = load_workbook(output_excel_filename)
        writer = pd.ExcelWriter(output_excel_filename, engine='openpyxl')
        writer.book = book
    except FileNotFoundError:
        # If the file does not exist, create a new one
        writer = pd.ExcelWriter(output_excel_filename, engine='openpyxl')
    
    # Save the DataFrame to Excel
    df.to_excel(writer, sheet_name=table_name, index=False)
    writer.save()
    writer.close()

    # Adjust the column width of the Excel sheet
    book = load_workbook(output_excel_filename)
    adjust_column_width_from_df(book[table_name])
    book.save(output_excel_filename)
    book.close()

    print(f"Data for the {table_name} table has been saved to Excel")
    
    # Also save to CSV
    df.to_csv(output_csv_filename, index=False, encoding='utf-8-sig')
    print(f"Data for the {table_name} table has been saved to CSV")
